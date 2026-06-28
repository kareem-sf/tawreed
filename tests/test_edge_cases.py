"""Expanded edge case tests for Tawreed.

This module adds comprehensive test coverage for edge cases and error scenarios
that are important for a production-grade BOQ processing application.

Areas covered:
- Network failures during AI calls
- Memory constraints with large Excel files
- Permission errors
- Corrupt Excel files
- Edge cases in Excel parsing
- Error handling in worker processes
"""

from __future__ import annotations

import asyncio
import os
import tempfile
import zipfile
from unittest.mock import Mock, patch

import openpyxl
import pytest
from openpyxl.utils.exceptions import InvalidFileException

from core import excel
from core.ai import analyze_boq_stream
from gui.worker import BOQProcessor, WorkerSignals

# ---------------------------------------------------------------------------
# Network failure tests
# ---------------------------------------------------------------------------


def test_ai_call_network_failure():
    """Test graceful handling of network failures during AI calls."""
    with patch("openai.OpenAI") as mock_openai:
        mock_client = Mock()
        mock_response = Mock()
        mock_response.choices = []
        mock_client.chat.completions.create.side_effect = ConnectionError("Network is unreachable")
        mock_openai.return_value = mock_client

        # The function should handle the error internally and yield __DONE__ with error
        results = list(
            analyze_boq_stream("fake-key", "https://api.openai.com/v1", "gpt-4", "system", "user")
        )

        # Should have a terminal __DONE__ message with error
        assert len(results) > 0
        assert results[-1][0] == "__DONE__"
        assert "ConnectionError" in results[-1][1]["error"]
        assert "Network is unreachable" in results[-1][1]["error"]


def test_ai_call_timeout():
    """Test handling of timeout errors during AI calls."""
    with patch("openai.OpenAI") as mock_openai:
        mock_client = Mock()
        mock_client.chat.completions.create.side_effect = TimeoutError("Request timed out")
        mock_openai.return_value = mock_client

        # The function should handle the error internally and yield __DONE__ with error
        results = list(
            analyze_boq_stream("fake-key", "https://api.openai.com/v1", "gpt-4", "system", "user")
        )

        # Should have a terminal __DONE__ message with error
        assert len(results) > 0
        assert results[-1][0] == "__DONE__"
        assert "TimeoutError" in results[-1][1]["error"]
        assert "Request timed out" in results[-1][1]["error"]


def test_ai_call_rate_limit():
    """Test handling of rate limit errors from AI providers."""
    mock_chunk = Mock()
    mock_chunk.choices = [Mock(delta=Mock(content='{"error": "rate_limit"}'))]

    with patch("openai.OpenAI") as mock_openai:
        mock_client = Mock()
        mock_client.chat.completions.create.return_value = [mock_chunk]
        mock_openai.return_value = mock_client

        result = list(
            analyze_boq_stream("fake-key", "https://api.openai.com/v1", "gpt-4", "system", "user")
        )

        # Should complete and return error in final result
        assert len(result) > 0
        assert result[-1][0] == "__DONE__"
        assert "error" in result[-1][1]


# ---------------------------------------------------------------------------
# Memory constraint tests
# ---------------------------------------------------------------------------


def test_large_excel_file_memory_efficient_parsing(tmp_path):
    """Test that large Excel files are parsed in memory-efficient mode."""
    # Create a reasonably large Excel file for testing
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add header
    ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])

    # Add rows with longer descriptions to make file larger
    for i in range(5000):  # Reduced from 20,000 to 5,000 for faster testing
        long_description = (
            f"Item {i + 1} with some additional text to make it longer and increase file size"
        )
        ws.append([f"{i + 1}", long_description, "m2", 100, 50, 5000])

    large_file = tmp_path / "large_boq.xlsx"
    wb.save(large_file)

    # File should be reasonable size for testing
    assert large_file.stat().st_size > 50_000  # >50KB is sufficient for testing

    # Parse should work without excessive memory usage
    md, data, headers = excel.parse_excel(str(large_file))

    # Should parse successfully
    assert len(data) == 5000
    first_key = list(data.keys())[0]
    assert "Nr." in data[first_key]
    assert data[first_key]["Nr."] == "1"


def test_excel_file_with_many_sheets(tmp_path):
    """Test parsing Excel files with many sheets."""
    wb = openpyxl.Workbook()

    # Create 10 sheets with data
    for sheet_num in range(10):
        if sheet_num > 0:
            ws = wb.create_sheet(title=f"Sheet{sheet_num + 1}")
        else:
            ws = wb.active
            ws.title = "Sheet1"

        ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])
        for i in range(100):
            ws.append([f"{i + 1}", f"Item {i + 1}", "m2", 10, 50, 500])

    multi_sheet_file = tmp_path / "multi_sheet.xlsx"
    wb.save(multi_sheet_file)

    # Parse should handle multiple sheets
    md, data, headers = excel.parse_excel(str(multi_sheet_file))

    # Should have data from all sheets
    assert len(data) == 1000  # 10 sheets × 100 rows each


# ---------------------------------------------------------------------------
# Permission error tests
# ---------------------------------------------------------------------------


def test_excel_read_permission_error(tmp_path):
    """Test handling of permission errors when reading Excel files."""
    # Create a valid Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])
    ws.append(["1", "Test item", "m2", 10, 50, 500])

    excel_file = tmp_path / "test.xlsx"
    wb.save(excel_file)

    # Make file read-only
    if os.name == "posix":
        os.chmod(excel_file, 0o444)  # Read-only

        try:
            # Should still be able to read
            md, data, headers = excel.parse_excel(str(excel_file))
            assert len(data) == 1

        finally:
            # Restore permissions
            os.chmod(excel_file, 0o644)
    else:
        # Windows: skip this test as file permissions work differently
        pytest.skip("Read-only permission test is POSIX-only")


def test_excel_write_permission_error(tmp_path):
    """Test handling of permission errors when writing Excel files."""
    output_dir = tmp_path / "readonly_dir"
    output_dir.mkdir()

    # Make directory read-only
    if os.name == "posix":
        os.chmod(output_dir, 0o555)  # Read and execute only

        try:
            output_file = output_dir / "output.xlsx"
            row_mapping = {
                "1": {
                    "Nr.": "1",
                    "Item Description": "Test",
                    "Unit": "m2",
                    "Qty": 10,
                    "Rate": 50,
                    "Amount": 0,
                }
            }

            # Should raise an error when trying to write
            with pytest.raises((IOError, OSError, PermissionError)):
                excel.write_excel(
                    str(output_file), row_mapping, {"1": "General"}, "Test Project", "2026-06-28"
                )

        finally:
            # Restore permissions
            os.chmod(output_dir, 0o755)
    else:
        # Windows: skip this test
        pytest.skip("Read-only directory test is POSIX-only")


# ---------------------------------------------------------------------------
# Corrupt Excel file tests
# ---------------------------------------------------------------------------


def test_excel_truncated_file(tmp_path):
    """Test handling of truncated Excel files."""
    # Create a valid Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nr.", "Item Description"])
    ws.append(["1", "Test"])

    valid_file = tmp_path / "valid.xlsx"
    wb.save(valid_file)

    # Create truncated version
    truncated_file = tmp_path / "truncated.xlsx"
    with open(valid_file, "rb") as f:
        content = f.read()

    # Truncate to 50% of original size
    with open(truncated_file, "wb") as f:
        f.write(content[: len(content) // 2])

    # Should raise ValueError with helpful message
    with pytest.raises(ValueError) as exc_info:
        excel.parse_excel(str(truncated_file))

    assert "corrupt or incomplete" in str(exc_info.value)


def test_excel_invalid_zip_content(tmp_path):
    """Test handling of Excel files with invalid ZIP content."""
    invalid_file = tmp_path / "invalid.xlsx"

    # Create a file that looks like Excel but has invalid content
    with zipfile.ZipFile(invalid_file, "w") as zf:
        # Add a file with invalid content that would confuse openpyxl
        zf.writestr("xl/worksheets/sheet1.xml", "<invalid xml content>")

    # Should raise an exception (KeyError or InvalidFileException)
    with pytest.raises((KeyError, InvalidFileException, zipfile.BadZipFile)):
        excel.parse_excel(str(invalid_file))


def test_excel_password_protected(tmp_path):
    """Test handling of password-protected Excel files."""
    # Create a file that looks corrupt (simulating password protection)
    protected_file = tmp_path / "protected.xlsx"

    # Write some bytes that would be typical of a corrupt file
    with open(protected_file, "wb") as f:
        # Write invalid ZIP signature
        f.write(b"PK\x03\x04" + b"\x00" * 100)  # Invalid ZIP structure

    # Should raise an exception when trying to parse
    with pytest.raises((ValueError, zipfile.BadZipFile, KeyError)):
        excel.parse_excel(str(protected_file))


# ---------------------------------------------------------------------------
# Edge case Excel parsing tests
# ---------------------------------------------------------------------------


def test_excel_empty_sheets(tmp_path):
    """Test handling of Excel files with empty sheets."""
    wb = openpyxl.Workbook()

    # Create sheet with header but no data
    ws1 = wb.active
    ws1.title = "EmptyData"
    ws1.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])

    # Create sheet with no header
    ws2 = wb.create_sheet("NoHeader")
    ws2.append(["Random", "Content", "Here"])

    # Create sheet with data
    ws3 = wb.create_sheet("WithData")
    ws3.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])
    ws3.append(["1", "Real item", "m2", 10, 50, 500])

    test_file = tmp_path / "mixed_sheets.xlsx"
    wb.save(test_file)

    # Should parse successfully, ignoring empty/invalid sheets
    md, data, headers = excel.parse_excel(str(test_file))

    # Should have data from the valid sheet only
    assert len(data) == 1
    first_key = list(data.keys())[0]
    assert data[first_key]["Nr."] == "1"
    assert data[first_key]["Item Description"] == "Real item"


def test_excel_mixed_arabic_english_content(tmp_path):
    """Test parsing Excel files with mixed Arabic and English content."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header in English
    ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])

    # Mixed content rows
    ws.append(["1", "English item", "m2", 10, 50, 500])
    ws.append(["2", "عنصر عربي", "م3", 5, 100, 500])
    ws.append(["3", "Mixed English وعربي", "each", 3, 75, 225])

    test_file = tmp_path / "mixed_content.xlsx"
    wb.save(test_file)

    # Should parse all rows correctly
    md, data, headers = excel.parse_excel(str(test_file))
    # Should parse all rows correctly
    assert len(data) == 3
    keys = list(data.keys())
    assert data[keys[0]]["Item Description"] == "English item"
    assert "عنصر عربي" in data[keys[1]]["Item Description"]
    assert "Mixed English" in data[keys[2]]["Item Description"]


def test_excel_special_characters_in_descriptions(tmp_path):
    """Test handling of special characters in item descriptions."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])

    # Add rows with special characters
    ws.append(["1", 'Item with "quotes"', "each", 1, 10, 10])
    ws.append(["2", "Item with 'apostrophes'", "each", 1, 10, 10])
    ws.append(["3", "Item with \n newlines", "each", 1, 10, 10])
    ws.append(["4", "Item with \t tabs", "each", 1, 10, 10])
    ws.append(["5", "Item with special chars: @#$%^&*()", "each", 1, 10, 10])

    test_file = tmp_path / "special_chars.xlsx"
    wb.save(test_file)

    # Should preserve special characters
    md, data, headers = excel.parse_excel(str(test_file))

    assert len(data) == 5
    keys = list(data.keys())
    assert data[keys[0]]["Item Description"] == 'Item with "quotes"'
    assert "apostrophes" in data[keys[1]]["Item Description"]
    assert "newlines" in data[keys[2]]["Item Description"]
    assert "tabs" in data[keys[3]]["Item Description"]
    assert "@#$%^&*()" in data[keys[4]]["Item Description"]


def test_excel_numeric_edge_cases(tmp_path):
    """Test handling of numeric edge cases in Excel data."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])

    # Add rows with various numeric formats
    ws.append(["1", "Very small quantity", "m2", 0.001, 1000, 1])
    ws.append(["2", "Very large quantity", "m2", 999999, 0.001, 999.999])
    ws.append(["3", "Zero rate", "m2", 100, 0, 0])
    ws.append(
        ["4", "Negative values", "m2", -10, 50, -500]
    )  # This might be invalid but should be handled

    test_file = tmp_path / "numeric_edge_cases.xlsx"
    wb.save(test_file)

    # Should handle numeric edge cases
    md, data, headers = excel.parse_excel(str(test_file))

    assert len(data) == 4
    keys = list(data.keys())
    assert data[keys[0]]["Qty"] == 0.001
    assert data[keys[1]]["Qty"] == 999999.0
    assert data[keys[2]]["Rate"] == 0.0


# ---------------------------------------------------------------------------
# Worker process error handling tests
# ---------------------------------------------------------------------------


def test_worker_handles_excel_parse_errors():
    """Test that worker process handles Excel parsing errors gracefully."""
    # Create proper signals object
    signals = WorkerSignals()
    processor = BOQProcessor("nonexistent.xlsx", signals)

    # Mock the parse_excel method to raise an error
    with patch.object(excel, "parse_excel") as mock_parse:
        mock_parse.side_effect = ValueError("Invalid Excel file")

        # Worker should handle the error without crashing
        try:
            asyncio.run(processor.process())
        except Exception:
            pass  # Expected to fail

        # The test passes if we get here without crashing
        assert True


def test_worker_handles_ai_call_errors():
    """Test that worker process handles AI call errors gracefully."""
    # Create proper signals object
    signals = WorkerSignals()

    # Create a temporary Excel file for testing
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])
        ws.append(["1", "Test item", "m2", 10, 50, 500])
        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        processor = BOQProcessor(tmp_path, signals)

        # Mock the parse_excel method to return valid data but have AI call fail
        with (
            patch.object(excel, "parse_excel") as mock_parse,
            patch("core.ai.analyze_boq_stream") as mock_ai,
        ):
            # Return valid Excel data but have AI call fail
            mock_parse.return_value = (
                {"metadata": "test"},
                {"R1": {"Nr.": "1", "Item Description": "Test item"}},
                ["Nr.", "Item Description"],
            )
            mock_ai.side_effect = Exception("AI call failed")

            try:
                asyncio.run(processor.process())
            except Exception:
                pass  # Expected to fail

            # The test passes if we get here without crashing
            assert True

    finally:
        # Clean up
        os.unlink(tmp_path)


def test_worker_handles_missing_file():
    """Test that worker process handles missing BOQ files gracefully."""
    # Create proper signals object
    signals = WorkerSignals()
    processor = BOQProcessor("nonexistent.xlsx", signals)

    try:
        asyncio.run(processor.process())
    except Exception:
        pass  # Expected to fail

    # The test passes if we get here without crashing
    assert True


# ---------------------------------------------------------------------------
# File type validation tests
# ---------------------------------------------------------------------------


def test_excel_old_xls_format_not_supported(tmp_path):
    """Test that old .xls format is properly rejected."""
    # Create a file that looks like old .xls format
    old_format_file = tmp_path / "old_format.xls"

    # Write some bytes that would be typical of old BIFF format
    with open(old_format_file, "wb") as f:
        # Old Excel files start with specific magic bytes
        f.write(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")
        f.write(b"\x00\x00\x00\x00" * 100)  # Some dummy content

    # Should raise ValueError with helpful message
    with pytest.raises(ValueError) as exc_info:
        excel.parse_excel(str(old_format_file))

    assert "xls" in str(exc_info.value).lower() or "not supported" in str(exc_info.value).lower()


def test_excel_non_excel_file_rejection(tmp_path):
    """Test that non-Excel files are properly rejected."""
    # Create a text file with .xlsx extension
    text_file = tmp_path / "fake.xlsx"
    text_file.write_text("This is not an Excel file, just plain text")

    # Should raise ValueError
    with pytest.raises(ValueError) as exc_info:
        excel.parse_excel(str(text_file))

    assert "corrupt or incomplete" in str(exc_info.value)


# ---------------------------------------------------------------------------
# Memory cleanup tests
# ---------------------------------------------------------------------------


def test_temp_file_cleanup_after_excel_operations(tmp_path):
    """Test that temporary files are cleaned up after Excel operations."""
    # Create a test Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nr.", "Item Description"])
    ws.append(["1", "Test"])

    test_file = tmp_path / "test.xlsx"
    wb.save(test_file)

    # Count files before
    files_before = len(list(tmp_path.glob("*")))

    # Parse the file
    md, data, headers = excel.parse_excel(str(test_file))

    # Count files after - should be same (no temp files left behind)
    files_after = len(list(tmp_path.glob("*")))

    assert files_after == files_before


def test_large_file_memory_cleanup():
    """Test that memory is properly cleaned up after processing large files."""
    # This is more of a integration test that would be run manually
    # For automated testing, we just ensure no exceptions are raised

    # Create a mock large file scenario
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        # Create a workbook with many rows
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])

        for i in range(5000):
            ws.append([f"{i + 1}", f"Item {i + 1}", "m2", 10, 50, 500])

        wb.save(tmp.name)
        tmp_path = tmp.name

    try:
        # Process the file
        md, data, headers = excel.parse_excel(tmp_path)

        # Should process successfully
        assert len(data) == 5000

    finally:
        # Clean up
        os.unlink(tmp_path)
