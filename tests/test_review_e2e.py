from __future__ import annotations

import asyncio

import openpyxl

from core import db
from core.processing_pipeline import BOQProcessingPipeline


class Signal:
    def __init__(self) -> None:
        self.values: list[object] = []

    def emit(self, value: object) -> None:
        self.values.append(value)


class Signals:
    def __init__(self) -> None:
        self.log = Signal()
        self.review_ready = Signal()
        self.progress = Signal()
        self.finished = Signal()
        self.error = Signal()


def test_parse_review_export_and_history_end_to_end(isolated_tawreed_dir):
    db.init_db()
    db.save_settings(
        {
            "provider": "OpenAI",
            "api_key": "test-key",
            "model": "test-model",
            "base_url": "https://api.openai.com/v1",
            "language": "en",
            "theme": "dark",
        }
    )
    source = isolated_tawreed_dir / "Pilot BOQ.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Nr.", "Item Description", "Unit", "Qty", "Rate", "Amount"])
    worksheet.append(["1", "Reinforced concrete", "m3", 10, 20, 200])
    worksheet.append(["2", "Block wall", "m2", 30, 4, 120])
    workbook.save(source)
    workbook.close()

    def classify(*_args):
        return {
            "project_name": "Ignored batch metadata",
            "date": "2026-07-10",
            "items": {"R1": "Concrete Works", "R2": "Masonry"},
        }

    signals = Signals()
    pipeline = BOQProcessingPipeline(
        str(source),
        signals,
        run_analysis_fn=classify,
        storage=db,
    )

    asyncio.run(pipeline.process())
    assert len(signals.review_ready.values) == 1, signals.error.values
    assert not signals.finished.values
    assert db.get_history() == []

    approval = signals.review_ready.values[0]
    assert not hasattr(approval, "items")
    assert approval.summary.total_items == 2
    asyncio.run(pipeline.approve_and_export(approval.token))

    assert len(signals.finished.values) == 1
    output = signals.finished.values[0]
    exported = openpyxl.load_workbook(output, data_only=False)
    try:
        master = exported["Master"]
        headers = [cell.value for cell in master[1]]
        package_column = headers.index("Package") + 1
        id_column = headers.index("Global ID") + 1
        assignments = {
            master.cell(row=row, column=id_column).value: master.cell(
                row=row,
                column=package_column,
            ).value
            for row in range(2, master.max_row + 1)
        }
    finally:
        exported.close()

    assert assignments == {"R1": "Concrete Works", "R2": "Masonry"}
    history = db.get_history()
    assert len(history) == 1
    assert history[0]["output_path"] == output
