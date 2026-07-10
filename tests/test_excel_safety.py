"""Safety and streaming regression tests for the Excel boundary."""

from __future__ import annotations

import openpyxl

from core import excel


def test_forced_streaming_path_handles_merged_header_and_sparse_tail(monkeypatch, tmp_path):
    source = tmp_path / "streaming.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("B1:C1")
    ws["A1"] = "Nr."
    ws["B1"] = "Item Description"
    ws["D1"] = "Unit"
    ws["E1"] = "Qty"
    ws["F1"] = "Rate"
    ws.append(["1", "Concrete work", None, "m3", 10, 50])
    ws["A4"] = "Section footer without a description"
    wb.save(source)
    wb.close()

    # Force a small fixture through the production >10 MB branch.
    monkeypatch.setattr(excel, "LARGE_FILE_THRESHOLD", 0)
    _markdown, rows, _headers = excel.parse_excel(str(source))

    assert list(rows) == ["R1"]
    assert rows["R1"]["Item Description"] == "Concrete work"
    assert rows["R1"]["Qty"] == 10.0


def test_forced_streaming_path_handles_stacked_header(monkeypatch, tmp_path):
    source = tmp_path / "streaming-stacked.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ITEM", "ITEM", "ITEM", "ITEM", "UNIT RATE", "AMOUNT"])
    ws.append(["NO.", "DESCRIPTION", "UNIT", "QTY.", "IN L.E.", "IN L.E."])
    ws.append([1, "Concrete work", "m3", 10, 50, 500])
    wb.save(source)
    wb.close()

    monkeypatch.setattr(excel, "LARGE_FILE_THRESHOLD", 0)
    _markdown, rows, _headers = excel.parse_excel(str(source))

    assert list(rows) == ["R1"]
    assert rows["R1"]["Item Description"] == "Concrete work"
    assert rows["R1"]["Qty"] == 10.0


def test_external_text_is_neutralized_but_trusted_amount_formula_remains(tmp_path):
    output = tmp_path / "safe.xlsx"
    row_mapping = {
        "R1": {
            "Nr.": "+1",
            "Item Description": '=HYPERLINK("https://example.invalid","click")',
            "Unit": "@unit",
            "Qty": 2,
            "Rate": 3,
            "sheet_name": "=Source",
        }
    }
    excel.write_excel(
        str(output),
        row_mapping,
        {"R1": '=WEBSERVICE("https://example.invalid")'},
        '=HYPERLINK("https://example.invalid","project")',
        "+2026-01-01",
    )

    wb = openpyxl.load_workbook(output, data_only=False)
    try:
        assert wb["Cover"]["B3"].data_type != "f"
        assert wb["Cover"]["B3"].value.startswith("'=")
        assert wb["Cover"]["B4"].value.startswith("'+")

        package_sheet = next(ws for ws in wb.worksheets if ws.title.startswith("Pkg - "))
        assert package_sheet["B2"].data_type != "f"
        assert package_sheet["B2"].value.startswith("'=")
        assert package_sheet["F2"].data_type == "f"

        master = wb["Master"]
        assert master["G2"].value.startswith("'=")
        assert master["H2"].value.startswith("'=")
    finally:
        wb.close()

    assert not list(tmp_path.glob(".tawreed-*.xlsx"))


def test_master_sheet_preserves_every_reviewed_global_id(tmp_path):
    output = tmp_path / "audit.xlsx"
    rows = {
        "R1": {
            "Nr.": "1",
            "Item Description": "Concrete",
            "Unit": "m3",
            "Qty": 1,
            "Rate": 2,
            "Amount": 2,
        },
        "R2": {
            "Nr.": "2",
            "Item Description": "Block wall",
            "Unit": "m2",
            "Qty": 3,
            "Rate": 4,
            "Amount": 12,
        },
    }
    excel.write_excel(
        str(output),
        rows,
        {"R1": "Concrete Works", "R2": "Masonry"},
        "Project",
        "2026-07-10",
    )

    workbook = openpyxl.load_workbook(output, data_only=False)
    try:
        master = workbook["Master"]
        headers = [cell.value for cell in master[1]]
        global_id_column = headers.index("Global ID") + 1
        exported_ids = {
            master.cell(row=row, column=global_id_column).value
            for row in range(2, master.max_row + 1)
        }
        assert exported_ids == {"R1", "R2"}
    finally:
        workbook.close()
