"""Test Excel cover sheet internationalization."""

from __future__ import annotations

import os
import tempfile

import openpyxl

from core.excel import write_excel
from core.i18n import get_i18n
from core.metadata import __version__


def test_cover_sheet_uses_english_when_no_i18n():
    """When no i18n context is provided, cover sheet uses English."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_output.xlsx")

        # Minimal data for a valid Excel
        row_mapping = {
            "R1": {
                "Nr.": "1",
                "Item Description": "Test item",
                "Unit": "m²",
                "Qty": 10,
                "Rate": 100,
                "Amount": 1000,
            }
        }
        item_categories = {"R1": "General"}
        project_name = "Test Project"
        date = "2026-06-28"

        # Call without i18n parameter
        write_excel(output_path, row_mapping, item_categories, project_name, date)

        # Verify the cover sheet
        wb = openpyxl.load_workbook(output_path)
        ws_cover = wb["Cover"]

        assert ws_cover["A1"].value == "Tawreed"
        assert ws_cover["B1"].value == "BOQ Work-Package Extractor"
        assert ws_cover["A3"].value == "Project Name"
        assert ws_cover["A4"].value == "Date"
        assert ws_cover["A5"].value == "Application"
        assert ws_cover["B5"].value == f"Tawreed BOQ Processor v{__version__}"


def test_cover_sheet_uses_arabic_when_i18n_provided():
    """When Arabic i18n context is provided, cover sheet uses Arabic."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_output_ar.xlsx")

        # Minimal data for a valid Excel
        row_mapping = {
            "R1": {
                "Nr.": "1",
                "Item Description": "Test item",
                "Unit": "m²",
                "Qty": 10,
                "Rate": 100,
                "Amount": 1000,
            }
        }
        item_categories = {"R1": "General"}
        project_name = "Test Project"
        date = "2026-06-28"

        # Create Arabic i18n context
        i18n = get_i18n()
        i18n.set_language("ar")

        # Call with Arabic i18n parameter
        write_excel(output_path, row_mapping, item_categories, project_name, date, i18n=i18n)

        # Verify the cover sheet
        wb = openpyxl.load_workbook(output_path)
        ws_cover = wb["Cover"]

        assert ws_cover["A1"].value == "توريد"
        assert ws_cover["B1"].value == "مستخرج حزم العمل من جدول الكميات"
        assert ws_cover["A3"].value == "اسم المشروع"
        assert ws_cover["A4"].value == "التاريخ"
        assert ws_cover["A5"].value == "التطبيق"
        assert ws_cover["B5"].value == f"معالج جدول الكميات توريد v{__version__}"


def test_cover_sheet_uses_english_when_i18n_is_english():
    """When English i18n context is provided, cover sheet uses English."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_output_en.xlsx")

        # Minimal data for a valid Excel
        row_mapping = {
            "R1": {
                "Nr.": "1",
                "Item Description": "Test item",
                "Unit": "m²",
                "Qty": 10,
                "Rate": 100,
                "Amount": 1000,
            }
        }
        item_categories = {"R1": "General"}
        project_name = "Test Project"
        date = "2026-06-28"

        # Create English i18n context
        i18n = get_i18n()
        i18n.set_language("en")

        # Call with English i18n parameter
        write_excel(output_path, row_mapping, item_categories, project_name, date, i18n=i18n)

        # Verify the cover sheet
        wb = openpyxl.load_workbook(output_path)
        ws_cover = wb["Cover"]

        assert ws_cover["A1"].value == "Tawreed"
        assert ws_cover["B1"].value == "BOQ Work-Package Extractor"
        assert ws_cover["A3"].value == "Project Name"
        assert ws_cover["A4"].value == "Date"
        assert ws_cover["A5"].value == "Application"
        assert ws_cover["B5"].value == f"Tawreed BOQ Processor v{__version__}"


def test_cover_sheet_project_name_and_date_preserved():
    """Project name and date are correctly written to cover sheet regardless of language."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_output.xlsx")

        # Minimal data for a valid Excel
        row_mapping = {
            "R1": {
                "Nr.": "1",
                "Item Description": "Test item",
                "Unit": "m²",
                "Qty": 10,
                "Rate": 100,
                "Amount": 1000,
            }
        }
        item_categories = {"R1": "General"}
        project_name = "My Test Project"
        date = "2026-06-28"

        # Create Arabic i18n context
        i18n = get_i18n()
        i18n.set_language("ar")

        # Call with Arabic i18n parameter
        write_excel(output_path, row_mapping, item_categories, project_name, date, i18n=i18n)

        # Verify the cover sheet
        wb = openpyxl.load_workbook(output_path)
        ws_cover = wb["Cover"]

        # Project name and date should be preserved as provided
        assert ws_cover["B3"].value == "My Test Project"
        assert ws_cover["B4"].value == "2026-06-28"


def test_cover_sheet_empty_project_name_shows_placeholder():
    """When project name is empty, it shows the placeholder."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_output.xlsx")

        # Minimal data for a valid Excel
        row_mapping = {
            "R1": {
                "Nr.": "1",
                "Item Description": "Test item",
                "Unit": "m²",
                "Qty": 10,
                "Rate": 100,
                "Amount": 1000,
            }
        }
        item_categories = {"R1": "General"}
        project_name = ""  # Empty project name
        date = "2026-06-28"

        # Create English i18n context
        i18n = get_i18n()
        i18n.set_language("en")

        # Call with English i18n parameter
        write_excel(output_path, row_mapping, item_categories, project_name, date, i18n=i18n)

        # Verify the cover sheet
        wb = openpyxl.load_workbook(output_path)
        ws_cover = wb["Cover"]

        # Empty project name should show placeholder
        assert ws_cover["B3"].value == "—"
        # And should have yellow highlight
        assert ws_cover["B3"].fill.start_color.index == "FFFEF3C7"


def test_cover_sheet_empty_date_shows_placeholder():
    """When date is empty, it shows the placeholder."""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_output.xlsx")

        # Minimal data for a valid Excel
        row_mapping = {
            "R1": {
                "Nr.": "1",
                "Item Description": "Test item",
                "Unit": "m²",
                "Qty": 10,
                "Rate": 100,
                "Amount": 1000,
            }
        }
        item_categories = {"R1": "General"}
        project_name = "Test Project"
        date = ""  # Empty date

        # Create Arabic i18n context
        i18n = get_i18n()
        i18n.set_language("ar")

        # Call with Arabic i18n parameter
        write_excel(output_path, row_mapping, item_categories, project_name, date, i18n=i18n)

        # Verify the cover sheet
        wb = openpyxl.load_workbook(output_path)
        ws_cover = wb["Cover"]

        # Empty date should show placeholder
        assert ws_cover["B4"].value == "—"
        # And should have yellow highlight
        assert ws_cover["B4"].fill.start_color.index == "FFFEF3C7"
